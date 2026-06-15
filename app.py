"""
Kit Production Planner — Interface Streamlit
Copyright © Thibaut MINE
"""

import hashlib
import json
import math
import io
from datetime import date
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from kit_planner import (
    Params, Kit, Composant, BOMLine,
    planifier, calcul_rapport_composants,
    ecrire_plan_prod, ecrire_rapport_composants,
    score_priorite,
)

# ─────────────────────────────────────────────
# HELPERS — CONVERSION
# ─────────────────────────────────────────────

_KITS_COLS  = ["Ref Kit", "Stock Actuel", "Demande Mensuelle", "MOQ", "Multiple", "Freeze"]
_BOM_COLS   = ["Ref Kit", "Ref Composant", "Qté Requise"]
_COMP_COLS  = ["Ref Composant", "Stock Disponible"]
_PARAMS_DEFAULT = {
    "capacite_semaine":      500,
    "periode_stock_mois":    3,
    "max_lancements":        20,
    "forcer_prod_sous_moq":  False,
}


def _data_hash(kits_df, bom_df, comp_df, params_dict) -> str:
    payload = json.dumps({
        "k": kits_df.to_json(),
        "b": bom_df.to_json(),
        "c": comp_df.to_json(),
        "p": json.dumps(params_dict, sort_keys=True),
    }, sort_keys=True)
    return hashlib.md5(payload.encode()).hexdigest()


def wb_to_dfs(wb):
    """Workbook → (params_dict, kits_df, bom_df, comp_df)."""
    ws_p = wb["Paramètres"]
    params = {
        "capacite_semaine":     int(ws_p["B2"].value or 500),
        "periode_stock_mois":   int(ws_p["B3"].value or 3),
        "max_lancements":       int(ws_p["B4"].value or 20),
        "forcer_prod_sous_moq": False,
    }

    ws_d = wb["Données"]
    # Structure : A=Ref  B=Stock  C=Demande  D=MOQ  E=Multiple  F=Freeze
    kits_rows = []
    for row in ws_d.iter_rows(min_row=2, values_only=True):
        ref = str(row[0]).strip() if row[0] else None
        if not ref or ref == "None":
            continue
        kits_rows.append({
            "Ref Kit":           ref,
            "Stock Actuel":      int(row[1] or 0),
            "Demande Mensuelle": int(row[2] or 0),
            "MOQ":               int(row[3] or 0),
            "Multiple":          int(row[4] or 1),
            "Freeze":            str(row[5] or "").strip().upper() == "X",
        })
    kits_df = pd.DataFrame(kits_rows) if kits_rows else pd.DataFrame(columns=_KITS_COLS)

    ws_b = wb["BOM"]
    bom_rows = []
    for row in ws_b.iter_rows(min_row=2, values_only=True):
        rk = str(row[0]).strip() if row[0] else None
        rc = str(row[1]).strip() if row[1] else None
        q  = int(row[2] or 0)
        if not rk or not rc or q <= 0:
            continue
        bom_rows.append({"Ref Kit": rk, "Ref Composant": rc, "Qté Requise": q})
    bom_df = pd.DataFrame(bom_rows) if bom_rows else pd.DataFrame(columns=_BOM_COLS)

    ws_c = wb["Composants"]
    comp_rows = []
    for row in ws_c.iter_rows(min_row=2, values_only=True):
        ref = str(row[0]).strip() if row[0] else None
        if not ref or ref == "None":
            continue
        comp_rows.append({"Ref Composant": ref, "Stock Disponible": int(row[1] or 0)})
    comp_df = pd.DataFrame(comp_rows) if comp_rows else pd.DataFrame(columns=_COMP_COLS)

    return params, kits_df, bom_df, comp_df


def dfs_to_objects(params_dict, kits_df, bom_df, comp_df):
    """DataFrames → objets kit_planner pour la planification."""
    params = Params(
        capacite_semaine      = int(params_dict.get("capacite_semaine", 500)),
        periode_stock_mois    = int(params_dict.get("periode_stock_mois", 3)),
        max_lancements        = int(params_dict.get("max_lancements", 20)),
        forcer_prod_sous_moq  = bool(params_dict.get("forcer_prod_sous_moq", False)),
    )

    kits = {}
    for _, row in kits_df.iterrows():
        ref = str(row.get("Ref Kit", "")).strip()
        if not ref or ref in ("nan", "None", ""):
            continue
        dem   = int(row.get("Demande Mensuelle", 0) or 0)
        conso = max(1, math.ceil(dem / 4)) if dem > 0 else 0
        stock = int(row.get("Stock Actuel", 0) or 0)
        moq_r = int(row.get("MOQ", 0) or 0)
        mult_r = int(row.get("Multiple", 1) or 1)
        kit = Kit(
            ref=ref,
            stock_actuel=stock,
            demande_mensuelle=dem,
            moq=moq_r if moq_r > 0 else params.moq_global,
            multiple=mult_r if mult_r > 1 else 1,
            freeze=bool(row.get("Freeze", False)),
            objectif_stock=dem * params.periode_stock_mois,
            conso_hebdo=conso,
        )
        kit.stock_previsionnel = kit.stock_actuel - kit.conso_hebdo
        kits[ref] = kit

    bom = {}
    for _, row in bom_df.iterrows():
        rk = str(row.get("Ref Kit", "")).strip()
        rc = str(row.get("Ref Composant", "")).strip()
        q  = int(row.get("Qté Requise", 0) or 0)
        if not rk or not rc or q <= 0 or rk == "nan" or rc == "nan":
            continue
        bom.setdefault(rk, []).append(BOMLine(rk, rc, q))

    composants = {}
    for _, row in comp_df.iterrows():
        ref = str(row.get("Ref Composant", "")).strip()
        if not ref or ref in ("nan", "None", ""):
            continue
        stock = int(row.get("Stock Disponible", 0) or 0)
        composants[ref] = Composant(ref=ref, stock=stock, stock_initial=stock)

    return params, kits, bom, composants


# ─────────────────────────────────────────────
# HELPERS — EXCEL GENERATION
# ─────────────────────────────────────────────

@st.cache_data
def generate_template() -> bytes:
    """Génère un fichier Excel template avec données d'exemple."""
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Paramètres"
    data_params = [
        ("Paramètre", "Valeur"),
        ("Capacité semaine (unités)", 500),
        ("Période stock objectif (mois)", 3),
        ("Max lancements / semaine", 20),
    ]
    for r, (a, b) in enumerate(data_params, 1):
        ws.cell(row=r, column=1, value=a)
        ws.cell(row=r, column=2, value=b)
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 12

    ws_d = wb.create_sheet("Données")
    # Colonnes : A=Ref Kit  B=Stock Actuel  C=Demande Mensuelle  D=MOQ  E=Multiple  F=Freeze(X)
    headers_d = ["Ref Kit", "Stock Actuel", "Demande Mensuelle", "MOQ", "Multiple", "Freeze (X)"]
    for c, h in enumerate(headers_d, 1):
        ws_d.cell(row=1, column=c, value=h).font = Font(bold=True)
    for r, row in enumerate([
        ("KIT-001", 50,  100, 20, 5, ""),
        ("KIT-002", 200, 80,  10, 1, ""),
        ("KIT-003", 0,   60,  15, 5, "X"),
        ("KIT-004", 30,  40,  0,  1, ""),
    ], 2):
        for c, val in enumerate(row, 1):
            ws_d.cell(row=r, column=c, value=val)

    ws_b = wb.create_sheet("BOM")
    for c, h in enumerate(["Ref Kit", "Ref Composant", "Qté Requise"], 1):
        ws_b.cell(row=1, column=c, value=h).font = Font(bold=True)
    for r, row in enumerate([
        ("KIT-001", "COMP-A", 2),
        ("KIT-001", "COMP-B", 1),
        ("KIT-002", "COMP-A", 1),
        ("KIT-002", "COMP-C", 3),
        ("KIT-003", "COMP-B", 2),
        ("KIT-003", "COMP-D", 1),
    ], 2):
        for c, val in enumerate(row, 1):
            ws_b.cell(row=r, column=c, value=val)

    ws_c = wb.create_sheet("Composants")
    for c, h in enumerate(["Ref Composant", "Stock Disponible"], 1):
        ws_c.cell(row=1, column=c, value=h).font = Font(bold=True)
    for r, row in enumerate([
        ("COMP-A", 500),
        ("COMP-B", 200),
        ("COMP-C", 50),
        ("COMP-D", 300),
    ], 2):
        for c, val in enumerate(row, 1):
            ws_c.cell(row=r, column=c, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_data_only_excel(params_dict, kits_df, bom_df, comp_df) -> bytes:
    """Export des 4 onglets de données uniquement, sans plan de production."""
    wb = openpyxl.Workbook()
    ws_p = wb.active
    ws_p.title = "Paramètres"
    for r, (a, b) in enumerate([
        ("Paramètre", "Valeur"),
        ("Capacité semaine (unités)", params_dict["capacite_semaine"]),
        ("Période stock objectif (mois)", params_dict["periode_stock_mois"]),
        ("Max lancements / semaine", params_dict["max_lancements"]),
    ], 1):
        ws_p.cell(row=r, column=1, value=a)
        ws_p.cell(row=r, column=2, value=b)
    ws_p.column_dimensions["A"].width = 35
    ws_p.column_dimensions["B"].width = 12

    ws_d = wb.create_sheet("Données")
    for c, h in enumerate(["Ref Kit", "Stock Actuel", "Demande Mensuelle", "MOQ", "Multiple", "Freeze"], 1):
        ws_d.cell(row=1, column=c, value=h).font = Font(bold=True)
    for r, (_, row) in enumerate(kits_df.iterrows(), 2):
        for c, val in enumerate([
            str(row.get("Ref Kit", "")),
            int(row.get("Stock Actuel", 0) or 0),
            int(row.get("Demande Mensuelle", 0) or 0),
            int(row.get("MOQ", 0) or 0),
            int(row.get("Multiple", 1) or 1),
            "X" if row.get("Freeze") else "",
        ], 1):
            ws_d.cell(row=r, column=c, value=val)

    ws_b = wb.create_sheet("BOM")
    for c, h in enumerate(["Ref Kit", "Ref Composant", "Qté Requise"], 1):
        ws_b.cell(row=1, column=c, value=h).font = Font(bold=True)
    for r, (_, row) in enumerate(bom_df.iterrows(), 2):
        ws_b.cell(row=r, column=1, value=str(row.get("Ref Kit", "")))
        ws_b.cell(row=r, column=2, value=str(row.get("Ref Composant", "")))
        ws_b.cell(row=r, column=3, value=int(row.get("Qté Requise", 0) or 0))

    ws_c = wb.create_sheet("Composants")
    for c, h in enumerate(["Ref Composant", "Stock Disponible"], 1):
        ws_c.cell(row=1, column=c, value=h).font = Font(bold=True)
    for r, (_, row) in enumerate(comp_df.iterrows(), 2):
        ws_c.cell(row=r, column=1, value=str(row.get("Ref Composant", "")))
        ws_c.cell(row=r, column=2, value=int(row.get("Stock Disponible", 0) or 0))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_output_excel(params_dict, kits_df, bom_df, comp_df, kits_planned, rapport_comps) -> bytes:
    """Génère l'Excel de sortie complet depuis les données courantes + résultats."""
    wb = openpyxl.Workbook()

    # Paramètres
    ws_p = wb.active
    ws_p.title = "Paramètres"
    for r, (a, b) in enumerate([
        ("Paramètre", "Valeur"),
        ("Capacité semaine (unités)", params_dict["capacite_semaine"]),
        ("Période stock objectif (mois)", params_dict["periode_stock_mois"]),
        ("Max lancements / semaine",  params_dict["max_lancements"]),
    ], 1):
        ws_p.cell(row=r, column=1, value=a)
        ws_p.cell(row=r, column=2, value=b)

    # Données
    ws_d = wb.create_sheet("Données")
    for c, h in enumerate(["Ref Kit", "Stock Actuel", "Demande Mensuelle", "MOQ", "Multiple", "Freeze"], 1):
        ws_d.cell(row=1, column=c, value=h).font = Font(bold=True)
    for r, (_, row) in enumerate(kits_df.iterrows(), 2):
        vals = [
            str(row.get("Ref Kit", "")),
            int(row.get("Stock Actuel", 0) or 0),
            int(row.get("Demande Mensuelle", 0) or 0),
            int(row.get("MOQ", 0) or 0),
            int(row.get("Multiple", 1) or 1),
            "X" if row.get("Freeze") else "",
        ]
        for c, val in enumerate(vals, 1):
            ws_d.cell(row=r, column=c, value=val)

    # BOM
    ws_b = wb.create_sheet("BOM")
    for c, h in enumerate(["Ref Kit", "Ref Composant", "Qté Requise"], 1):
        ws_b.cell(row=1, column=c, value=h).font = Font(bold=True)
    for r, (_, row) in enumerate(bom_df.iterrows(), 2):
        ws_b.cell(row=r, column=1, value=str(row.get("Ref Kit", "")))
        ws_b.cell(row=r, column=2, value=str(row.get("Ref Composant", "")))
        ws_b.cell(row=r, column=3, value=int(row.get("Qté Requise", 0) or 0))

    # Composants
    ws_c = wb.create_sheet("Composants")
    for c, h in enumerate(["Ref Composant", "Stock Disponible"], 1):
        ws_c.cell(row=1, column=c, value=h).font = Font(bold=True)
    for r, (_, row) in enumerate(comp_df.iterrows(), 2):
        ws_c.cell(row=r, column=1, value=str(row.get("Ref Composant", "")))
        ws_c.cell(row=r, column=2, value=int(row.get("Stock Disponible", 0) or 0))

    ecrire_plan_prod(wb, kits_planned)
    ecrire_rapport_composants(wb, rapport_comps)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────
# CONFIG STREAMLIT
# ─────────────────────────────────────────────

st.set_page_config(
    page_title="Kit Planner",
    page_icon="⚙️",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
    .main { background-color: #f8f9fb; }
    .block-container { padding-top: 1.2rem; }
    h1 { color: #1F3864; font-size: 1.6rem; }
    h2, h3 { color: #2E4057; }
    .stMetric > div { background: white; border-radius: 8px; padding: 10px; border: 1px solid #e0e0e0; }
    .status-ok     { color: #276221; background: #C6EFCE; padding: 2px 8px; border-radius: 4px; font-size: 0.82em; }
    .status-bloque { color: #9C0006; background: #FFCCCC; padding: 2px 8px; border-radius: 4px; font-size: 0.82em; }
    .status-gele   { color: #555;    background: #DDDDDD; padding: 2px 8px; border-radius: 4px; font-size: 0.82em; }
    .status-warn   { color: #7d4800; background: #FFEB9C; padding: 2px 8px; border-radius: 4px; font-size: 0.82em; }
    div[data-testid="stDataFrameResizable"] { border-radius: 6px; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────

with st.sidebar:
    st.markdown("## ⚙️ Kit Planner")
    st.caption("Planification de production hebdomadaire")
    st.divider()

    # Template download (toujours visible)
    st.markdown("**Fichier template**")
    st.download_button(
        label="📥 Télécharger le template Excel",
        data=generate_template(),
        file_name="kit_planner_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        help="Télécharge un fichier exemple avec la structure attendue",
    )

    st.divider()

    uploaded = st.file_uploader(
        "📂 Charger un fichier Excel",
        type=["xlsx"],
        help="Onglets requis : Données, BOM, Composants, Paramètres",
    )

    col_vierge = st.columns(1)[0]
    demarrer_vierge = col_vierge.button(
        "🆕 Démarrer vierge",
        use_container_width=True,
        help="Initialise un projet vide sans fichier",
    )

    st.divider()

    # ── Paramètres globaux ──
    st.markdown("**⚙️ Paramètres**")
    if "params_dict" not in st.session_state:
        st.session_state["params_dict"] = _PARAMS_DEFAULT.copy()

    p = st.session_state["params_dict"]
    capa  = st.number_input("Capacité semaine (unités)", value=int(p["capacite_semaine"]),   min_value=1, step=10)
    per   = st.number_input("Période stock (mois)",      value=int(p["periode_stock_mois"]), min_value=1)
    maxl  = st.number_input("Max lancements / semaine",  value=int(p["max_lancements"]),     min_value=1)
    force = st.checkbox(
        "Forcer prod. hors MOQ si rupture",
        value=bool(p.get("forcer_prod_sous_moq", False)),
        help=(
            "Si coché : les références en rupture critique (stock < 0 ou < 1 semaine) "
            "peuvent être produites même si la quantité disponible est inférieure au MOQ. "
            "Le statut affiché sera 'Planifié < MOQ'."
        ),
    )
    st.session_state["params_dict"] = {
        "capacite_semaine":     capa,
        "periode_stock_mois":   per,
        "max_lancements":       maxl,
        "forcer_prod_sous_moq": force,
    }

    st.divider()
    st.markdown("**Légende statuts**")
    st.markdown('<span class="status-ok">Planifié</span>', unsafe_allow_html=True)
    st.markdown('<span class="status-warn">Sous objectif</span>', unsafe_allow_html=True)
    st.markdown(
        '<span style="color:#7d4200;background:#FFD580;padding:2px 8px;border-radius:4px;font-size:0.82em;">'
        'Planifié &lt; MOQ</span>',
        unsafe_allow_html=True,
    )
    st.markdown('<span class="status-bloque">Bloqué composants</span>', unsafe_allow_html=True)
    st.markdown('<span class="status-bloque">MOQ non atteignable</span>', unsafe_allow_html=True)
    st.markdown('<span class="status-gele">Capacité épuisée</span>', unsafe_allow_html=True)
    st.markdown('<span class="status-gele">Max lancements atteint</span>', unsafe_allow_html=True)
    st.markdown('<span class="status-gele">Gelé / Inactif / Stock OK</span>', unsafe_allow_html=True)

    st.divider()
    st.markdown(
        '<p style="font-size:0.72em; color:#888; text-align:center; margin:0;">'
        '© Thibaut MINE — Tous droits réservés</p>',
        unsafe_allow_html=True,
    )


# ─────────────────────────────────────────────
# INITIALISATION SESSION STATE
# ─────────────────────────────────────────────

def _init_empty():
    st.session_state["kits_df"] = pd.DataFrame(columns=_KITS_COLS)
    st.session_state["bom_df"]  = pd.DataFrame(columns=_BOM_COLS)
    st.session_state["comp_df"] = pd.DataFrame(columns=_COMP_COLS)
    st.session_state.pop("results", None)
    st.session_state["last_file_key"] = "__vierge__"


# Démarrer vierge
if demarrer_vierge:
    _init_empty()
    st.session_state["params_dict"] = _PARAMS_DEFAULT.copy()
    st.toast("Projet vierge initialisé", icon="🆕")

# Chargement fichier
if uploaded:
    file_key = f"{uploaded.name}_{uploaded.size}"
    if st.session_state.get("last_file_key") != file_key:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(uploaded.getvalue()))
            params_loaded, kits_df, bom_df, comp_df = wb_to_dfs(wb)
            st.session_state["params_dict"] = params_loaded
            st.session_state["kits_df"]     = kits_df
            st.session_state["bom_df"]      = bom_df
            st.session_state["comp_df"]     = comp_df
            st.session_state.pop("results", None)
            st.session_state["last_file_key"] = file_key
        except Exception as e:
            st.error(f"Erreur de lecture du fichier : {e}")
            st.stop()

# Aucune donnée chargée
if "kits_df" not in st.session_state:
    st.title("⚙️ Kit Planner — Planification de Production")
    st.info("👈 Télécharge le template, remplis-le et charge-le ici — ou démarre avec un projet vierge.")
    st.stop()


# ─────────────────────────────────────────────
# MAIN — ONGLETS
# ─────────────────────────────────────────────

st.title("⚙️ Kit Planner — Planification de Production")

nb_kits = len(st.session_state["kits_df"])
nb_bom  = len(st.session_state["bom_df"])
nb_comp = len(st.session_state["comp_df"])
st.caption(f"**{nb_kits}** références · **{nb_bom}** lignes BOM · **{nb_comp}** composants")

tab_refs, tab_bom, tab_comp, tab_plan, tab_export, tab_regles = st.tabs([
    "📊 Références",
    "🔗 BOM",
    "📦 Composants",
    "▶️ Plan de production",
    "💾 Export",
    "📖 Règles d'allocation",
])


# ─── TAB 1 : Références ─────────────────────

with tab_refs:
    st.markdown(
        "Modifie directement dans le tableau. "
        "**Stock & demande** : à mettre à jour ici ou via l'Excel rechargé. "
        "**MOQ = 0** → pas de minimum de lancement."
    )

    kits_edited = st.data_editor(
        st.session_state["kits_df"],
        num_rows="dynamic",
        use_container_width=True,
        hide_index=True,
        column_config={
            "Ref Kit":           st.column_config.TextColumn("Ref Kit", width="medium"),
            "Stock Actuel":      st.column_config.NumberColumn("Stock Actuel",  min_value=0, format="%d"),
            "Demande Mensuelle": st.column_config.NumberColumn("Demande/mois",  min_value=0, format="%d"),
            "MOQ":               st.column_config.NumberColumn("MOQ",           min_value=0, format="%d", help="0 = pas de minimum"),
            "Multiple":          st.column_config.NumberColumn("Multiple",      min_value=1, format="%d", help="Arrondi au multiple supérieur"),
            "Freeze":            st.column_config.CheckboxColumn("Freeze", help="Coché = exclu de la planification"),
        },
        key="editor_kits",
    )
    st.session_state["kits_df"] = kits_edited
    st.caption(f"{len(kits_edited)} références — Ajoute des lignes avec le **+** en bas du tableau")


# ─── TAB 2 : BOM ────────────────────────────

with tab_bom:
    st.markdown(
        "Liste tous les composants requis par référence kit. "
        "**Qté Requise** = quantité de composant par unité de kit produite."
    )

    # Résumé par kit (pour info)
    if not st.session_state["bom_df"].empty:
        bom_summary = (
            st.session_state["bom_df"]
            .groupby("Ref Kit")["Ref Composant"]
            .count()
            .reset_index()
            .rename(columns={"Ref Composant": "Nb composants"})
        )
        with st.expander(f"Résumé BOM ({len(st.session_state['bom_df'])} lignes, {bom_summary['Ref Kit'].nunique()} kits)"):
            st.dataframe(bom_summary, use_container_width=True, hide_index=True)

    # Éditeur BOM
    kit_refs  = sorted(st.session_state["kits_df"]["Ref Kit"].dropna().unique().tolist()) if not st.session_state["kits_df"].empty else []
    comp_refs = sorted(st.session_state["comp_df"]["Ref Composant"].dropna().unique().tolist()) if not st.session_state["comp_df"].empty else []

    col_cfg_bom = {
        "Qté Requise": st.column_config.NumberColumn("Qté Requise", min_value=1, format="%d"),
    }
    if kit_refs:
        col_cfg_bom["Ref Kit"] = st.column_config.SelectboxColumn(
            "Ref Kit", options=kit_refs, width="medium", required=False,
        )
    else:
        col_cfg_bom["Ref Kit"] = st.column_config.TextColumn("Ref Kit", width="medium")
    if comp_refs:
        col_cfg_bom["Ref Composant"] = st.column_config.SelectboxColumn(
            "Ref Composant", options=comp_refs, width="medium", required=False,
        )
    else:
        col_cfg_bom["Ref Composant"] = st.column_config.TextColumn("Ref Composant", width="medium")

    bom_edited = st.data_editor(
        st.session_state["bom_df"],
        num_rows="dynamic",
        use_container_width=True,
        hide_index=True,
        column_config=col_cfg_bom,
        key="editor_bom",
    )
    st.session_state["bom_df"] = bom_edited
    st.caption("Ajoute des lignes avec le **+** en bas du tableau · Une ligne par composant par kit")

    # Kits sans BOM + refs orphelines
    if not st.session_state["kits_df"].empty and not bom_edited.empty:
        kits_avec_bom = set(bom_edited["Ref Kit"].dropna().unique())
        kits_sans_bom = [r for r in kit_refs if r not in kits_avec_bom]
        if kits_sans_bom:
            st.info(f"ℹ️ Kits sans BOM (pas de contrainte composants) : `{'`, `'.join(kits_sans_bom)}`")
        orphans = [r for r in kits_avec_bom if r and r not in kit_refs]
        if orphans:
            st.warning(f"⚠️ Lignes BOM référençant des kits absents de l'onglet Références : `{'`, `'.join(sorted(orphans))}`")


# ─── TAB 3 : Composants ─────────────────────

with tab_comp:
    st.markdown(
        "Stocks de composants disponibles. "
        "**Seules les quantités** sont à mettre à jour régulièrement — les références peuvent être gérées ici."
    )

    comp_edited = st.data_editor(
        st.session_state["comp_df"],
        num_rows="dynamic",
        use_container_width=True,
        hide_index=True,
        column_config={
            "Ref Composant":   st.column_config.TextColumn("Ref Composant", width="medium"),
            "Stock Disponible": st.column_config.NumberColumn("Stock Disponible", min_value=0, format="%d"),
        },
        key="editor_comp",
    )
    st.session_state["comp_df"] = comp_edited
    st.caption(f"{len(comp_edited)} composants — Ajoute des lignes avec le **+** en bas du tableau")

    # Composants utilisés en BOM mais absents du stock
    if not st.session_state["bom_df"].empty and not comp_edited.empty:
        comps_bom = set(st.session_state["bom_df"]["Ref Composant"].dropna().unique())
        comps_stock = set(comp_edited["Ref Composant"].dropna().unique())
        manquants = comps_bom - comps_stock
        if manquants:
            st.warning(
                f"⚠️ Composants présents en BOM mais absents du stock "
                f"(seront bloquants) : `{'`, `'.join(sorted(manquants))}`"
            )


# ─── TAB 4 : Plan de production ─────────────

with tab_plan:
    col_btn, col_sp = st.columns([1, 3])
    lancer = col_btn.button("▶️ Lancer la planification", type="primary", use_container_width=True)

    if lancer:
        if st.session_state["kits_df"].empty:
            st.error("Aucune référence définie — va dans l'onglet Références pour en ajouter.")
        else:
            with st.spinner("Calcul en cours..."):
                params, kits, bom, composants = dfs_to_objects(
                    st.session_state["params_dict"],
                    st.session_state["kits_df"],
                    st.session_state["bom_df"],
                    st.session_state["comp_df"],
                )
                kits = planifier(kits, bom, composants, params)
                rapport_comps = calcul_rapport_composants(kits, bom, composants)

            st.session_state["results"] = {
                "kits":          kits,
                "rapport_comps": rapport_comps,
                "capa":          st.session_state["params_dict"]["capacite_semaine"],
                "data_hash":     _data_hash(
                    st.session_state["kits_df"],
                    st.session_state["bom_df"],
                    st.session_state["comp_df"],
                    st.session_state["params_dict"],
                ),
            }
            st.success("Planification terminée.")

    if "results" not in st.session_state:
        st.info("Lance la planification pour voir les résultats.")
    else:
        # ── Stale data warning ──
        current_hash = _data_hash(
            st.session_state["kits_df"],
            st.session_state["bom_df"],
            st.session_state["comp_df"],
            st.session_state["params_dict"],
        )
        if current_hash != st.session_state["results"].get("data_hash"):
            st.warning(
                "⚠️ Les données ont été modifiées depuis la dernière planification — "
                "relancez le calcul pour mettre à jour les résultats."
            )

        # ── Résultats ──
        res = st.session_state["results"]
        kits_res = res["kits"]
        rapport_comps = res["rapport_comps"]
        capa = res["capa"]

        semaine_n1 = date.today().isocalendar()[1] + 1
        st.caption(
            f"Plan semaine **S{semaine_n1}** (N+1) · "
            f"calculé le {date.today().strftime('%d/%m/%Y')}"
        )

        # Kits produits = tout ce qui a production_planifiee > 0
        planifies = [k for k in kits_res.values() if k.production_planifiee > 0]
        bloques   = [k for k in kits_res.values() if k.statut in ("Bloqué composants", "MOQ non atteignable")]
        geles     = [k for k in kits_res.values() if k.statut == "Gelé"]
        sous_obj  = [k for k in kits_res.values() if k.statut in ("Sous objectif", "Planifié < MOQ")]
        total_prod    = sum(k.production_planifiee for k in planifies)
        capa_utilisee = round(total_prod / capa * 100, 1) if capa else 0

        st.divider()
        m1, m2, m3, m4, m5 = st.columns(5)
        m1.metric("✅ Refs planifiées",  len(planifies))
        m2.metric("📦 Unités planifiées", f"{total_prod:,}")
        m3.metric("📊 Capacité utilisée", f"{capa_utilisee}%")
        m4.metric("🔴 Refs bloquées",    len(bloques))
        m5.metric("⚠️ Sous objectif",    len(sous_obj))

        st.divider()

        # Filtres
        f1, f2, f3 = st.columns([2, 2, 1])
        filtre_statut     = f1.multiselect("Filtrer par statut", options=sorted(set(k.statut for k in kits_res.values())), default=[])
        filtre_ref        = f2.text_input("Rechercher une référence", "")
        show_bloques_only = f3.checkbox("Bloqués seulement", False)

        rows = []
        for kit in sorted(kits_res.values(), key=lambda k: (score_priorite(k)[0], -score_priorite(k)[1], k.ref)):
            rows.append({
                "Ref Kit":     kit.ref,
                "Stock Act.":  kit.stock_actuel,
                "Conso/sem":   kit.conso_hebdo,
                "Stock Prév.": kit.stock_previsionnel,
                "Objectif":    kit.objectif_stock,
                "Besoin":      kit.besoin_prod,
                "Production":  kit.production_planifiee,
                "Stock Final": kit.stock_final,
                "MOQ":         kit.moq,
                "Multiple":    kit.multiple,
                "Statut":      kit.statut,
                "Blocage":     kit.raison_blocage,
            })

        df = pd.DataFrame(rows)
        if filtre_statut:
            df = df[df["Statut"].isin(filtre_statut)]
        if filtre_ref:
            df = df[df["Ref Kit"].str.contains(filtre_ref, case=False, na=False)]
        if show_bloques_only:
            df = df[df["Statut"].isin(["Bloqué composants", "MOQ non atteignable"])]

        def colorize(row):
            s = row["Statut"]
            if s == "Planifié":
                return ["background-color: #C6EFCE"] * len(row)
            if s == "Sous objectif":
                return ["background-color: #FFEB9C"] * len(row)
            if s == "Planifié < MOQ":
                return ["background-color: #FFD580"] * len(row)
            if s in ("Bloqué composants", "MOQ non atteignable"):
                return ["background-color: #FFCCCC"] * len(row)
            if s == "Gelé":
                return ["background-color: #EEEEEE; color: #888"] * len(row)
            return [""] * len(row)

        st.dataframe(
            df.style.apply(colorize, axis=1),
            use_container_width=True,
            height=480,
            hide_index=True,
            column_config={
                "Production":  st.column_config.NumberColumn(format="%d", help="Unités planifiées cette semaine"),
                "Stock Final": st.column_config.NumberColumn(format="%d"),
                "Blocage":     st.column_config.TextColumn(width="large"),
            },
        )
        st.caption(f"{len(df)} références affichées")

        # ── Composants critiques (sous-section) ──
        st.divider()
        with st.expander(f"🔧 Composants critiques ({len(rapport_comps)})", expanded=bool(rapport_comps)):
            if not rapport_comps:
                st.success("✅ Aucun composant limitant.")
            else:
                st.warning(f"{len(rapport_comps)} composants limitent la production")
                df_comp = pd.DataFrame(rapport_comps)

                def colorize_comp(row):
                    nb = row["Nb Kits Bloqués"]
                    if nb >= 5:
                        return ["background-color: #FF9999"] * len(row)
                    if nb >= 2:
                        return ["background-color: #FFCCCC"] * len(row)
                    return [""] * len(row)

                st.dataframe(
                    df_comp.style.apply(colorize_comp, axis=1),
                    use_container_width=True,
                    height=320,
                    hide_index=True,
                    column_config={
                        "Kits Bloqués (détail)":  st.column_config.TextColumn(width="large"),
                        "Qté Manquante Totale":   st.column_config.NumberColumn(format="%d"),
                        "Kits Débloqués si Appro": st.column_config.NumberColumn(
                            format="%d", help="Kits débloqués si composant approvisionné"
                        ),
                    },
                )
                st.markdown("**Top 5 à approvisionner :**")
                for _, r in df_comp.sort_values("Nb Kits Bloqués", ascending=False).head(5).iterrows():
                    st.markdown(
                        f"- **{r['Ref Composant']}** — {r['Nb Kits Bloqués']} kits bloqués "
                        f"| Manque : {r['Qté Manquante Totale']:,} | Débloque : {r['Kits Débloqués si Appro']} kits"
                    )


# ─── TAB 5 : Export ─────────────────────────

with tab_export:
    st.markdown("### Télécharger le fichier Excel complet")
    st.markdown(
        "Exporte **toutes les données** (Références, BOM, Composants, Paramètres) "
        "telles qu'éditées dans l'interface, plus les onglets Plan et Rapport générés par la planification."
    )

    has_results = "results" in st.session_state

    if not has_results:
        st.warning("Lance d'abord la planification (onglet **Plan de production**) pour inclure les résultats.")

    col_a, col_b = st.columns(2)

    with col_a:
        # Export données seules — 4 onglets, sans onglet Plan
        data_bytes = build_data_only_excel(
            st.session_state["params_dict"],
            st.session_state["kits_df"],
            st.session_state["bom_df"],
            st.session_state["comp_df"],
        )
        st.download_button(
            label="📄 Exporter les données (sans plan)",
            data=data_bytes,
            file_name="kit_planner_donnees.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            help="Sauvegarde Références + BOM + Composants + Paramètres pour réimportation",
        )

    with col_b:
        if has_results:
            res = st.session_state["results"]
            plan_bytes = build_output_excel(
                st.session_state["params_dict"],
                st.session_state["kits_df"],
                st.session_state["bom_df"],
                st.session_state["comp_df"],
                res["kits"],
                res["rapport_comps"],
            )
            st.download_button(
                label="📥 Exporter le plan complet (.xlsx)",
                data=plan_bytes,
                file_name="plan_production_kits.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )

    st.divider()
    if has_results:
        res = st.session_state["results"]
        kits_res    = res["kits"]
        rapport_res = res["rapport_comps"]
        capa_r      = res["capa"]
        semaine_n1_exp = date.today().isocalendar()[1] + 1
        plan_list   = [k for k in kits_res.values() if k.production_planifiee > 0]
        blq_list    = [k for k in kits_res.values() if k.statut in ("Bloqué composants", "MOQ non atteignable")]
        gel_list    = [k for k in kits_res.values() if k.statut == "Gelé"]
        sobj_list   = [k for k in kits_res.values() if k.statut == "Sous objectif"]
        moq_list    = [k for k in kits_res.values() if k.statut == "Planifié < MOQ"]
        tot         = sum(k.production_planifiee for k in plan_list)
        capa_pct    = round(tot / capa_r * 100, 1) if capa_r else 0
        st.markdown("### Résumé")
        st.code(
            f"PLAN DE PRODUCTION — SEMAINE S{semaine_n1_exp}\n"
            f"{'='*40}\n"
            f"Refs planifiées   : {len(plan_list)}\n"
            f"  dont sous obj.  : {len(sobj_list)}\n"
            f"  dont < MOQ      : {len(moq_list)}\n"
            f"Unités totales    : {tot:,}\n"
            f"Capacité utilisée : {capa_pct}%\n"
            f"Refs bloquées     : {len(blq_list)}\n"
            f"Refs gelées       : {len(gel_list)}\n"
            f"Composants crit.  : {len(rapport_res)}\n",
            language=None,
        )


# ─── TAB 6 : Règles d'allocation ────────────

with tab_regles:
    st.markdown("## Règles d'allocation de production")
    st.markdown(
        "Cette page décrit la logique complète appliquée par le moteur de planification "
        "lors de chaque exécution. Elle sert de référence pour interpréter les résultats "
        "et paramétrer correctement le système."
    )

    # ── 1. Vue d'ensemble ──
    with st.expander("1 — Vue d'ensemble du processus", expanded=True):
        st.markdown("""
Le moteur suit **5 étapes séquentielles** à chaque planification :

| Étape | Action |
|---|---|
| **1** | Calcul du stock prévisionnel pour chaque référence |
| **2** | Calcul du besoin de production |
| **3** | Scoring et tri par priorité |
| **4** | Allocation itérative (respect capacité, MOQ, composants) |
| **5** | Rapport des composants critiques |

> Le calcul est **hebdomadaire** : il planifie la production de la **semaine N+1**.
""")

    # ── 2. Stock prévisionnel ──
    with st.expander("2 — Calcul du stock prévisionnel"):
        st.markdown("""
Avant d'évaluer le besoin, le stock actuel est ajusté de la consommation prévue sur la semaine à venir.

```
Conso hebdo      = ceil(Demande mensuelle / 4)
Stock prévisionnel = Stock actuel − Conso hebdo
```

Si le **stock prévisionnel est négatif**, la référence est en **rupture** et obtient la priorité absolue.

> **Exemple :** Stock actuel = 45, Demande mensuelle = 200 → Conso hebdo = 50 → Stock prévisionnel = **−5** (rupture)
""")

    # ── 3. Calcul du besoin ──
    with st.expander("3 — Calcul du besoin de production"):
        st.markdown("""
Le **besoin** est la quantité à produire pour atteindre l'objectif de stock.

```
Objectif stock   = Demande mensuelle × Période stock (mois)
```

| Situation | Besoin calculé |
|---|---|
| Stock prévisionnel < 0 (rupture) | abs(Stock prév.) + Objectif stock |
| 0 ≤ Stock prév. < Objectif | Objectif − Stock prévisionnel |
| Stock prév. ≥ Objectif | 0 (pas de besoin) |

Les références **sans besoin** (stock ≥ objectif) ne sont pas soumises à l'algorithme d'allocation.
""")

    # ── 4. Priorité ──
    with st.expander("4 — Système de priorité (scoring)"):
        st.markdown("""
Chaque référence reçoit un **niveau de priorité** (1 = le plus urgent) et un **score** au sein de ce niveau.
Le tri est effectué **niveau croissant, puis score décroissant**.

| Niveau | Condition | Score de base |
|---|---|---|
| **1 — Rupture critique** | Stock prév. < 0 | 10 000 000 + abs(écart) × 1000 |
| **1 — Rupture totale** | Stock prév. = 0 | 5 000 000 |
| **1 — Stock < 1 semaine** | Stock prév. ≤ Conso hebdo | 2 000 000 |
| **1 — Stock < 2 semaines** | Stock prév. ≤ Conso hebdo × 2 | 1 000 000 |
| **2 — Stock < 50 % objectif** | Stock prév. < 50 % objectif | 100 000 × (1 − taux) |
| **2 — Stock < objectif** | Stock prév. < objectif | 10 000 × (1 − taux) |
| **3 — Stock OK** | Stock prév. ≥ objectif | Score faible, basé sur demande |

> La demande mensuelle (**volume**) sert de critère de départage à égalité de niveau : une référence à fort volume est prioritaire sur une référence à faible volume au sein du même niveau.
""")

    # ── 5. Contraintes d'allocation ──
    with st.expander("5 — Contraintes appliquées à chaque allocation"):
        st.markdown("""
Pour chaque référence traitée dans l'ordre de priorité, le moteur vérifie :

**A. Capacité disponible**
```
Production ≤ Capacité semaine restante
```
Si la capacité est épuisée, la référence est marquée `Capacité épuisée` et ignorée.

---

**B. Limite du nombre de lancements**
```
Nb références déjà planifiées < Max lancements / semaine
```
Si le plafond est atteint, la référence est marquée `Max lancements atteint`.

---

**C. Disponibilité composants (BOM)**
```
Pour chaque composant C requis :
    Qté disponible(C) / Qté requise par kit ≥ MOQ
```
Si un composant est manquant pour produire au moins le MOQ, la référence est marquée `Bloqué composants`.
Le détail du composant bloquant et de la quantité manquante est affiché dans la colonne **Blocage**.

---

**D. MOQ et Multiple**
```
Production cible = max(Besoin, MOQ)
Production cible = ceil(Production cible / Multiple) × Multiple
```
Si la `Production cible` ainsi calculée dépasse la capacité restante ou le stock composants disponible,
la référence est marquée `MOQ non atteignable` (le MOQ ne peut pas être respecté cette semaine).

---

**E. Option — Forcer production sous MOQ (rupture critique)**

Lorsque l'option **"Forcer prod. sous MOQ si rupture critique"** est activée dans la sidebar :
- Si une référence est en **rupture critique** (niveau 1) et que la production cible dépasse la capacité ou les composants disponibles,
  le moteur tente de produire la **quantité maximale possible** même si elle est inférieure au MOQ.
- La référence est alors marquée `Planifié < MOQ` et la raison précise la quantité forcée.
- Ce mécanisme ne s'applique **qu'aux ruptures critiques** — les niveaux 2 et 3 ne bénéficient pas de ce traitement.
""")

    # ── 6. Statuts possibles ──
    with st.expander("6 — Tableau des statuts"):
        st.markdown("""
| Statut | Couleur | Signification |
|---|---|---|
| **Planifié** | 🟢 Vert | Produit cette semaine, stock final ≥ objectif |
| **Sous objectif** | 🟡 Jaune clair | Produit cette semaine, mais stock final reste < objectif (capacité ou composants insuffisants pour couvrir l'objectif complet) |
| **Planifié < MOQ** | 🟠 Or | Produit en quantité inférieure au MOQ — uniquement pour ruptures critiques avec l'option "Forcer prod. sous MOQ" activée |
| **Bloqué composants** | 🔴 Rouge | Composant(s) insuffisant(s) pour atteindre le MOQ |
| **MOQ non atteignable** | 🔴 Rouge | Capacité ou stock composants insuffisant pour respecter le MOQ (option forçage désactivée ou niveau non critique) |
| **Capacité épuisée** | ⬜ Blanc | Capacité semaine consommée avant traitement de cette ref |
| **Max lancements atteint** | ⬜ Blanc | Plafond du nombre de références lancées atteint |
| **Stock OK** | ⬜ Blanc | Stock ≥ objectif, aucune production nécessaire |
| **Gelé** | ⬛ Gris | Référence marquée Freeze = X, exclue de tout calcul |
| **Inactif** | ⬛ Gris | Demande mensuelle = 0, référence sans activité |
""")

    # ── 7. Gestion composants ──
    with st.expander("7 — Gestion des stocks composants"):
        st.markdown("""
Les stocks composants sont **consommés au fur et à mesure** des allocations, dans l'ordre de priorité des kits.

- Un kit planifié en priorité 1 consomme ses composants **avant** les kits de priorité 2.
- Le stock disponible d'un composant = `Stock initial − Quantité déjà allouée aux kits précédents`.
- Si un composant n'est **pas référencé dans l'onglet Composants**, il est considéré **absent** → bloquant immédiat.
- Si un kit **n'a pas de BOM**, il est considéré sans contrainte composants (production libre sous réserve de capacité).

**Rapport composants critiques** : après planification, le rapport identifie pour chaque composant bloquant :
- Le nombre de kits qu'il bloque
- La quantité manquante totale
- Le nombre de kits supplémentaires qui seraient débloqués si ce composant était approvisionné
""")

    # ── 8. Paramètres ──
    with st.expander("8 — Paramètres de planification"):
        st.markdown("""
| Paramètre | Rôle | Effet si augmenté |
|---|---|---|
| **Capacité semaine** | Nombre total d'unités produisibles sur la semaine, toutes références confondues | Plus de kits planifiés |
| **Période stock (mois)** | Horizon de stock cible. Ex : 3 → objectif = 3 mois de demande | Objectifs plus élevés, plus de besoins calculés |
| **Max lancements / semaine** | Nombre maximum de références différentes lancées en production | Limite la dispersion des ordres |

**MOQ (Minimum Order Quantity)** et **Multiple** sont définis **par référence** dans l'onglet Références :
- `MOQ = 0` → pas de minimum (toute quantité ≥ 1 est acceptée)
- `Multiple = 5` → la quantité produite sera toujours un multiple de 5 (ex : 20, 25, 30…)
""")

    # ── 9. Conseils ──
    with st.expander("9 — Bonnes pratiques et conseils d'utilisation"):
        st.markdown("""
**Mise à jour hebdomadaire recommandée :**
1. Mettre à jour les **stocks actuels** et **demandes mensuelles** dans l'onglet Références (ou via l'Excel rechargé)
2. Mettre à jour les **stocks composants** dans l'onglet Composants
3. Vérifier les paramètres de la sidebar (capacité, période, max lancements)
4. Activer **"Forcer prod. sous MOQ si rupture critique"** si nécessaire (kits en rupture avec MOQ élevé)
5. Cliquer **Lancer la planification**
6. Analyser les composants critiques et planifier les approvisionnements
7. Exporter le plan Excel pour diffusion

**Recommandations :**
- Définir un **MOQ par référence** cohérent avec la taille de lot réelle en production
- Un `MOQ = 0` est équivalent à `MOQ = 1` (aucun minimum imposé)
- Utiliser le **Freeze** pour exclure temporairement une référence sans la supprimer (outillage indisponible, arrêt commercial…)
- Le paramètre **Max lancements** permet de limiter les changements de série : une valeur basse favorise les grandes séries, une valeur haute maximise la couverture de stock
- Les kits sans BOM ne sont jamais bloqués par les composants : ils consomment de la capacité mais pas de composants

**Limites connues :**
- La planification est **mono-semaine** : elle ne lisse pas la charge sur plusieurs semaines
- L'option "Forcer prod. sous MOQ" ne s'applique qu'aux ruptures de niveau 1 — les références en simple sous-couverture ne bénéficient pas de ce traitement
""")

    st.divider()
    st.markdown(
        '<p style="font-size:0.78em; color:#aaa; text-align:right;">© Thibaut MINE — Kit Planner</p>',
        unsafe_allow_html=True,
    )
