"""
Kit Production Planner
Copyright © Thibaut MINE
Planification hebdomadaire avec gestion BOM, MOQ/multiple par ref, freeze, criticité composants
"""

import math
import sys
from dataclasses import dataclass, field
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# DATA STRUCTURES
# ─────────────────────────────────────────────

@dataclass
class Params:
    capacite_semaine: int                   # B2 — capacité prod totale (unités/semaine)
    periode_stock_mois: int = 3             # B3 — objectif stock en mois
    max_lancements: int = 999               # B4 — nb max de références différentes/semaine
    moq_global: int = 1                     # interne — MOQ géré à la ref, fallback = 1
    forcer_prod_sous_moq: bool = False      # forcer la prod hors MOQ pour les ruptures critiques (niveau 1)


@dataclass
class Kit:
    ref: str
    stock_actuel: int
    demande_mensuelle: int
    moq: int                    # D — MOQ spécifique par référence (0 = pas de minimum)
    multiple: int               # E — multiple de lancement (ex: 5 → multiples de 5)
    freeze: bool                # F — "X" = exclu de la planification
    # Calculé
    objectif_stock: int = 0
    conso_hebdo: int = 0
    stock_previsionnel: int = 0
    besoin_prod: int = 0
    production_planifiee: int = 0
    stock_final: int = 0
    statut: str = ""
    raison_blocage: str = ""


@dataclass
class Composant:
    ref: str
    stock: int
    stock_initial: int = 0

    @property
    def disponible(self) -> int:
        return self.stock


@dataclass
class BOMLine:
    ref_kit: str
    ref_composant: str
    qte_requise: int


# ─────────────────────────────────────────────
# LECTURE EXCEL
# ─────────────────────────────────────────────

def load_params(wb) -> Params:
    ws = wb["Paramètres"]
    return Params(
        capacite_semaine   = int(ws["B2"].value or 500),
        periode_stock_mois = int(ws["B3"].value or 3),
        max_lancements     = int(ws["B4"].value or 999),
    )


def load_kits(wb, params: Params) -> dict[str, Kit]:
    """
    Structure Données (colonnes) :
    A=Ref Kit  B=Stock Actuel  C=Demande Mensuelle  D=MOQ  E=Multiple  F=Freeze(X)
    """
    ws = wb["Données"]
    kits = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        ref = str(row[0]).strip() if row[0] else None
        if not ref or ref == "None":
            continue
        dem          = int(row[2] or 0)
        conso_hebdo  = max(1, math.ceil(dem / 4)) if dem > 0 else 0
        objectif     = dem * params.periode_stock_mois
        stock        = int(row[1] or 0)
        moq_ref      = int(row[3] or 0)   # colonne D
        multiple_ref = int(row[4] or 1)   # colonne E
        freeze       = str(row[5] or "").strip().upper() == "X"  # colonne F

        kit = Kit(
            ref=ref,
            stock_actuel=stock,
            demande_mensuelle=dem,
            moq=moq_ref if moq_ref > 0 else params.moq_global,
            multiple=multiple_ref if multiple_ref > 1 else 1,
            freeze=freeze,
            objectif_stock=objectif,
            conso_hebdo=conso_hebdo,
        )
        kit.stock_previsionnel = kit.stock_actuel - kit.conso_hebdo
        kits[ref] = kit
    return kits


def load_bom(wb) -> dict[str, list[BOMLine]]:
    ws = wb["BOM"]
    bom: dict[str, list[BOMLine]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        ref_kit  = str(row[0]).strip() if row[0] else None
        ref_comp = str(row[1]).strip() if row[1] else None
        qte      = int(row[2] or 0)
        if not ref_kit or not ref_comp or qte <= 0:
            continue
        bom.setdefault(ref_kit, []).append(BOMLine(ref_kit, ref_comp, qte))
    return bom


def load_composants(wb) -> dict[str, Composant]:
    ws = wb["Composants"]
    comps = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        ref   = str(row[0]).strip() if row[0] else None
        stock = int(row[1] or 0)
        if not ref or ref == "None":
            continue
        comps[ref] = Composant(ref=ref, stock=stock, stock_initial=stock)
    return comps


# ─────────────────────────────────────────────
# LOGIQUE COMPOSANTS
# ─────────────────────────────────────────────

def max_kits_composants(ref_kit: str, bom: dict, composants: dict) -> int:
    """Retourne le nombre max de kits productibles selon le stock composants disponible."""
    if ref_kit not in bom:
        return 999999  # pas de BOM = pas de contrainte composants
    max_possible = 999999
    for line in bom[ref_kit]:
        comp = composants.get(line.ref_composant)
        if comp is None:
            return 0  # composant inconnu = bloquant immédiat
        if line.qte_requise <= 0:
            continue
        possible = comp.disponible // line.qte_requise
        max_possible = min(max_possible, possible)
    return max_possible


def allouer_composants(ref_kit: str, qte: int, bom: dict, composants: dict):
    """Déduit les composants consommés du stock après validation d'un lancement."""
    if ref_kit not in bom:
        return
    for line in bom[ref_kit]:
        comp = composants.get(line.ref_composant)
        if comp:
            comp.stock = max(0, comp.stock - line.qte_requise * qte)


# ─────────────────────────────────────────────
# AJUSTEMENT MOQ / MULTIPLE
# ─────────────────────────────────────────────

def ajuster_au_multiple(qte: int, multiple: int) -> int:
    if multiple <= 1:
        return qte
    return math.ceil(qte / multiple) * multiple


def appliquer_moq_multiple(qte: int, moq: int, multiple: int) -> int:
    qte = max(qte, moq)
    return ajuster_au_multiple(qte, multiple)


# ─────────────────────────────────────────────
# SCORE DE PRIORITÉ
# ─────────────────────────────────────────────

def score_priorite(kit: Kit) -> tuple[int, float]:
    """Retourne (niveau, score) — niveau 1 = urgence absolue, score décroissant = plus urgent."""
    stock = kit.stock_previsionnel
    dem   = kit.demande_mensuelle
    if dem <= 0:
        return (4, 0.0)
    if stock < 0:
        return (1, 10_000_000 + abs(stock) * 1000 + dem)
    elif stock == 0:
        return (1, 5_000_000 + dem)
    elif stock <= kit.conso_hebdo:
        return (1, 2_000_000 + (kit.conso_hebdo - stock) * 100 + dem)
    elif stock <= kit.conso_hebdo * 2:
        return (1, 1_000_000 + (kit.conso_hebdo * 2 - stock) * 50 + dem)
    elif stock < kit.objectif_stock:
        pct = stock / kit.objectif_stock if kit.objectif_stock > 0 else 1
        if pct < 0.5:
            return (2, 100_000 * (1 - pct) * (1 + dem / 100))
        else:
            return (2, 10_000 * (1 - pct) * (1 + dem / 1000))
    else:
        return (3, dem / 100 if dem > 100 else 0.0)


# ─────────────────────────────────────────────
# ALGORITHME DE PLANIFICATION
# ─────────────────────────────────────────────

def planifier(kits: dict[str, Kit], bom: dict, composants: dict, params: Params) -> dict[str, Kit]:
    capacite_restante = params.capacite_semaine
    nb_lancements     = 0

    # ── Calcul du besoin initial ──
    for kit in kits.values():
        if kit.freeze or kit.demande_mensuelle <= 0:
            kit.statut     = "Gelé" if kit.freeze else "Inactif"
            kit.stock_final = kit.stock_previsionnel
            continue
        stock = kit.stock_previsionnel
        if stock < 0:
            kit.besoin_prod = abs(stock) + kit.objectif_stock
        elif stock < kit.objectif_stock:
            kit.besoin_prod = kit.objectif_stock - stock
        else:
            kit.besoin_prod = 0

    # ── Tri par priorité : niveau croissant, score décroissant ──
    kits_actifs = [
        k for k in kits.values()
        if not k.freeze and k.demande_mensuelle > 0 and k.besoin_prod > 0
    ]
    kits_actifs.sort(key=lambda k: (score_priorite(k)[0], -score_priorite(k)[1]))

    # ── Allocation ──
    for kit in kits_actifs:

        if capacite_restante <= 0:
            kit.statut      = "Capacité épuisée"
            kit.stock_final = kit.stock_previsionnel
            continue

        if nb_lancements >= params.max_lancements:
            kit.statut      = "Max lancements atteint"
            kit.stock_final = kit.stock_previsionnel
            continue

        # Contrainte composants
        max_comp = max_kits_composants(kit.ref, bom, composants)
        if max_comp == 0:
            kit.statut      = "Bloqué composants"
            kit.raison_blocage = _detail_composants_manquants(kit.ref, kit.moq, bom, composants)
            kit.stock_final = kit.stock_previsionnel
            continue

        # Production brute = min(besoin, capacité restante, max composants)
        prod_brute = min(kit.besoin_prod, capacite_restante, max_comp)

        # Application MOQ + multiple
        prod_cible = appliquer_moq_multiple(prod_brute, kit.moq, kit.multiple)

        # Vérification : la prod cible est-elle compatible avec les ressources ?
        if prod_cible > max_comp or prod_cible > capacite_restante:
            # Fix 3 — prod forcée hors MOQ pour les ruptures critiques (niveau 1)
            if params.forcer_prod_sous_moq and score_priorite(kit)[0] == 1:
                prod_urgence = min(max_comp, capacite_restante, kit.besoin_prod)
                if prod_urgence > 0:
                    allouer_composants(kit.ref, prod_urgence, bom, composants)
                    kit.production_planifiee = prod_urgence
                    kit.stock_final          = kit.stock_previsionnel + prod_urgence
                    kit.statut               = "Planifié < MOQ"
                    kit.raison_blocage       = (
                        f"Prod. forcée : {prod_urgence} unités (MOQ={kit.moq} non respecté) "
                        f"— rupture critique"
                    )
                    capacite_restante -= prod_urgence
                    nb_lancements     += 1
                    continue

            kit.statut      = "MOQ non atteignable"
            kit.raison_blocage = (
                f"MOQ={kit.moq}, multiple={kit.multiple} → besoin {prod_cible} "
                f"| Dispo composants : {max_comp} | Capacité : {capacite_restante}"
            )
            kit.stock_final = kit.stock_previsionnel
            continue

        # Lancement validé — allocation composants
        allouer_composants(kit.ref, prod_cible, bom, composants)

        kit.production_planifiee = prod_cible
        kit.stock_final          = kit.stock_previsionnel + prod_cible
        capacite_restante        -= prod_cible
        nb_lancements            += 1

        # Fix 1 — Planifié mais stock final sous objectif (capacité ou comps insuffisants)
        if kit.stock_final >= kit.objectif_stock:
            kit.statut = "Planifié"
        else:
            kit.statut = "Sous objectif"

    # ── Kits sans besoin ou non traités ──
    for kit in kits.values():
        if kit.statut == "":
            kit.stock_final = kit.stock_previsionnel
            kit.statut      = "Stock OK"

    return kits


def _detail_composants_manquants(ref_kit: str, moq: int, bom: dict, composants: dict) -> str:
    if ref_kit not in bom:
        return "BOM absente"
    manquants = []
    for line in bom[ref_kit]:
        comp    = composants.get(line.ref_composant)
        besoin  = line.qte_requise * max(moq, 1)
        if comp is None:
            manquants.append(f"{line.ref_composant}: absent du stock")
        elif comp.disponible < besoin:
            manquants.append(
                f"{line.ref_composant}: besoin {besoin}, dispo {comp.disponible} "
                f"(manque {besoin - comp.disponible})"
            )
    return " | ".join(manquants) if manquants else ""


# ─────────────────────────────────────────────
# RAPPORT CRITICITÉ COMPOSANTS
# ─────────────────────────────────────────────

def calcul_rapport_composants(kits: dict[str, Kit], bom: dict, composants: dict) -> list[dict]:
    criticite: dict[str, dict] = {}

    for kit in kits.values():
        if kit.statut not in ("Bloqué composants", "MOQ non atteignable"):
            continue
        if kit.ref not in bom:
            continue
        for line in bom[kit.ref]:
            comp  = composants.get(line.ref_composant)
            dispo = comp.disponible if comp else 0
            besoin_moq = line.qte_requise * max(kit.moq, 1)
            if dispo >= besoin_moq and comp:
                continue
            manque = max(0, besoin_moq - dispo)
            if line.ref_composant not in criticite:
                criticite[line.ref_composant] = {
                    "ref_composant":        line.ref_composant,
                    "stock_dispo":          dispo,
                    "kits_bloques":         [],
                    "qte_manquante_totale": 0,
                    "kits_deblocables":     0,
                }
            criticite[line.ref_composant]["kits_bloques"].append(kit.ref)
            criticite[line.ref_composant]["qte_manquante_totale"] += manque
            if line.qte_requise > 0:
                criticite[line.ref_composant]["kits_deblocables"] += manque // line.qte_requise

    rows = []
    for c in criticite.values():
        rows.append({
            "Ref Composant":          c["ref_composant"],
            "Stock Disponible":       c["stock_dispo"],
            "Qté Manquante Totale":   c["qte_manquante_totale"],
            "Nb Kits Bloqués":        len(set(c["kits_bloques"])),
            "Kits Bloqués (détail)":  ", ".join(set(c["kits_bloques"])),
            "Kits Débloqués si Appro": c["kits_deblocables"],
        })
    rows.sort(key=lambda r: r["Nb Kits Bloqués"], reverse=True)
    return rows


# ─────────────────────────────────────────────
# ÉCRITURE EXCEL
# ─────────────────────────────────────────────

CLR_HEADER     = "1F3864"
CLR_PLANIFIE   = "C6EFCE"
CLR_BLOQUE     = "FFCCCC"
CLR_GELE       = "DDDDDD"
CLR_ATTENTION  = "FFEB9C"
CLR_URGENCE    = "FFD580"   # orange clair — Planifié < MOQ
CLR_RAPPORT_HDR = "2E4057"

THIN   = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _header_style(cell, bg=CLR_HEADER):
    cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BORDER


def _cell_style(cell, bg=None, bold=False, align="left", num_format=None):
    cell.font      = Font(bold=bold, name="Arial", size=9)
    if bg:
        cell.fill  = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = BORDER
    if num_format:
        cell.number_format = num_format


def ecrire_plan_prod(wb, kits: dict[str, Kit]):
    if "Plan de prod" in wb.sheetnames:
        del wb["Plan de prod"]
    ws = wb.create_sheet("Plan de prod", 0)

    headers    = ["Ref Kit", "Stock Actuel", "Conso Hebdo", "Stock Prévisionnel",
                  "Objectif Stock", "Besoin Prod", "Production Planifiée", "Stock Final",
                  "MOQ", "Multiple", "Statut", "Raison Blocage"]
    col_widths = [18, 13, 13, 18, 14, 12, 18, 12, 8, 10, 22, 55]

    ws.row_dimensions[1].height = 30
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        _header_style(ws.cell(row=1, column=i, value=h))
        ws.column_dimensions[get_column_letter(i)].width = w

    def sort_key(k):
        ordre = {"Planifié": 0, "Sous objectif": 1, "Planifié < MOQ": 2,
                 "Bloqué composants": 3, "MOQ non atteignable": 4,
                 "Capacité épuisée": 5, "Max lancements atteint": 6,
                 "Stock OK": 7, "Inactif": 8, "Gelé": 9}
        niv, score = score_priorite(k)
        return (ordre.get(k.statut, 99), niv, -score)

    for r, kit in enumerate(sorted(kits.values(), key=sort_key), 2):
        s  = kit.statut
        bg = (CLR_PLANIFIE  if s == "Planifié"        else
              CLR_ATTENTION  if s == "Sous objectif"   else
              CLR_URGENCE    if s == "Planifié < MOQ"  else
              CLR_BLOQUE     if s in ("Bloqué composants", "MOQ non atteignable") else
              CLR_GELE       if s == "Gelé"            else None)

        vals = [kit.ref, kit.stock_actuel, kit.conso_hebdo, kit.stock_previsionnel,
                kit.objectif_stock, kit.besoin_prod, kit.production_planifiee, kit.stock_final,
                kit.moq, kit.multiple, s, kit.raison_blocage]
        for c, val in enumerate(vals, 1):
            cell  = ws.cell(row=r, column=c, value=val)
            align = "right" if c in (2, 3, 4, 5, 6, 7, 8, 9, 10) else "left"
            _cell_style(cell, bg=bg, align=align, bold=(c == 7 and kit.production_planifiee > 0))

    nb_planifies = sum(1 for k in kits.values() if k.production_planifiee > 0)
    total_prod   = sum(k.production_planifiee for k in kits.values())
    r_recap      = len(kits) + 2
    ws.cell(row=r_recap, column=1,  value="TOTAL").font = Font(bold=True, name="Arial")
    ws.cell(row=r_recap, column=7,  value=total_prod).font = Font(bold=True, name="Arial")
    ws.cell(row=r_recap, column=11, value=f"{nb_planifies} références produites").font = Font(bold=True, name="Arial")

    ws.freeze_panes = "A2"


def ecrire_rapport_composants(wb, rapport: list[dict]):
    if "Rapport Composants" in wb.sheetnames:
        del wb["Rapport Composants"]
    ws = wb.create_sheet("Rapport Composants")

    if not rapport:
        ws["A1"] = "Aucun composant critique — tous les kits sont productibles."
        return

    headers    = list(rapport[0].keys())
    col_widths = [20, 16, 22, 16, 60, 22]

    ws.row_dimensions[1].height = 30
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        _header_style(ws.cell(row=1, column=i, value=h), bg=CLR_RAPPORT_HDR)
        ws.column_dimensions[get_column_letter(i)].width = w

    for r, row in enumerate(rapport, 2):
        nb  = row["Nb Kits Bloqués"]
        bg  = "FF9999" if nb >= 5 else ("FFCCCC" if nb >= 2 else None)
        for c, (key, val) in enumerate(row.items(), 1):
            align = "right" if c in (2, 3, 4, 6) else "left"
            _cell_style(ws.cell(row=r, column=c, value=val), bg=bg, align=align)

    ws.freeze_panes = "A2"


# ─────────────────────────────────────────────
# MAIN CLI
# ─────────────────────────────────────────────

def run(filepath: str):
    print(f"Chargement : {filepath}")
    wb = openpyxl.load_workbook(filepath)

    params     = load_params(wb)
    kits       = load_kits(wb, params)
    bom        = load_bom(wb)
    composants = load_composants(wb)

    print(f"  {len(kits)} kits | {len(bom)} refs BOM | {len(composants)} composants")
    print(f"  Capacité : {params.capacite_semaine} | Max lancements : {params.max_lancements}")

    kits      = planifier(kits, bom, composants, params)
    rapport   = calcul_rapport_composants(kits, bom, composants)

    produits  = [k for k in kits.values() if k.production_planifiee > 0]
    bloques   = [k for k in kits.values() if k.statut in ("Bloqué composants", "MOQ non atteignable")]
    print(f"\n Produits   : {len(produits)} refs | {sum(k.production_planifiee for k in produits)} unités")
    print(f" Bloqués    : {len(bloques)} refs")
    print(f" Composants critiques : {len(rapport)}")

    ecrire_plan_prod(wb, kits)
    ecrire_rapport_composants(wb, rapport)

    output_path = filepath.replace(".xlsx", "_planifie.xlsx")
    if output_path == filepath:
        output_path = filepath + "_planifie.xlsx"
    wb.save(output_path)
    print(f"\n Sauvegardé : {output_path}")
    return output_path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage : python kit_planner.py <fichier.xlsx>")
        sys.exit(1)
    run(sys.argv[1])
